import openpyxl


def get_data(sheetName):
    workbook = openpyxl.load_workbook("../excel/testdata.xlsx")
    sheet = workbook[sheetName]
    totalRows = sheet.max_row
    totalCols = sheet.max_column
    print("total cols are ", str(totalCols))
    print("total rows are ", str(totalRows))
    mainList = []

    for i in range(2, totalRows + 1):
        dataList = []
        for j in range(1, totalCols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
        print(mainList)
    return mainList
